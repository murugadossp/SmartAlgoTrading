"""Parse portfolio CSV (and optionally Excel) to list of holdings: symbol, quantity, avg_cost or value."""
import csv
import io
import warnings
from typing import Any

# Optional Excel: add openpyxl to requirements if needed
try:
    import openpyxl
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# Broker exports (e.g. Groww) often omit default style; openpyxl warns then applies its own. Harmless.
_OPENPYXL_STYLE_WARNING = "Workbook contains no default style"


def file_content_to_text(content: bytes, filename: str = "") -> str:
    """
    Convert uploaded file content to a single string for LLM consumption.
    CSV: decoded as UTF-8/latin-1. Excel: first sheet as CSV-like table (header + rows).
    Returns empty string if file type unsupported or sheet empty.
    """
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        try:
            return content.decode("utf-8-sig").strip()
        except UnicodeDecodeError:
            try:
                return content.decode("latin-1").strip()
            except Exception:
                return ""
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        if not HAS_EXCEL:
            return ""
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", message=_OPENPYXL_STYLE_WARNING)
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            if not ws:
                wb.close()
                return ""
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            return ""
        if not rows:
            return ""
        # Serialize as CSV (escape commas in cells)
        out = io.StringIO()
        writer = csv.writer(out)
        for row in rows:
            writer.writerow(["" if v is None else str(v).strip() for v in (row or [])])
        return out.getvalue().strip()
    return ""


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_").replace("-", "_")


def _detect_columns(row: dict[str, str], use_all_keys: bool = False) -> dict[str, str | None]:
    """Return map of role -> header key: symbol, stock_name, quantity, avg_cost, buy_value, closing_price, value, unrealized_pnl.
    value = current/closing value (not buy value). If use_all_keys=True, use all keys for detection (e.g. header row).
    """
    keys = [k for k in row.keys() if k] if use_all_keys else [k for k in row.keys() if k and row.get(k) is not None]
    out: dict[str, str | None] = {
        "symbol": None,
        "stock_name": None,
        "quantity": None,
        "avg_cost": None,
        "buy_value": None,
        "closing_price": None,
        "value": None,
        "unrealized_pnl": None,
    }
    # Prefer ticker/symbol columns for symbol; keep "Stock Name" / "Company" for stock_name
    sym_ticker_names = ("symbol", "symbol_id", "ticker", "scrip", "trading_symbol", "instrument", "script")
    stock_name_names = ("stock_name", "stock name", "company", "security", "name")
    qty_names = (
        "quantity", "qty", "qty.", "shares", "units", "balance_quantity", "balance_qty", "available_balance", "balance"
    )
    cost_names = (
        "avg_cost", "average_cost", "cost_price", "buy_price", "avg_price", "average_price", "purchase_price",
        "average_buy_price",
    )
    buy_value_names = ("buy_value", "buy value", "total_cost", "cost_value")
    closing_price_names = ("closing_price", "closing price", "last_price", "current_price", "last_traded_price")
    current_value_names = (
        "value", "current_value", "current value", "closing_value", "closing value", "market_value",
        "total_value", "valuation", "amount",
    )
    pnl_names = ("unrealised_pnl", "unrealised p&l", "unrealized_pnl", "unrealized pnl", "pnl", "profit_loss", "profit loss")
    for k in keys:
        n = _normalize_header(k)
        if n in sym_ticker_names or (n.endswith("_symbol") and "trading" in n) or n == "trading_symbol":
            out["symbol"] = k
        elif n in stock_name_names or n == "stock name" or n == "company":
            out["stock_name"] = k
        elif n in qty_names or "quantity" in n or "qty" in n:
            out["quantity"] = k
        elif n in cost_names or ("average" in n and "price" in n):
            out["avg_cost"] = k
        elif n in buy_value_names or (n == "buy value"):
            out["buy_value"] = k
        elif n in closing_price_names or (n == "closing price"):
            out["closing_price"] = k
        elif n in current_value_names or (n == "closing value") or (n == "current value"):
            # Prefer closing/current value; avoid assigning "buy value" here
            if "buy" not in n and "cost" not in n:
                out["value"] = k
        elif n in pnl_names or "unrealised" in n or "unrealized" in n:
            out["unrealized_pnl"] = k
    # Fallback: if no symbol yet, use first column that looks like symbol/ticker/stock name
    if not out["symbol"]:
        for k in keys:
            nn = _normalize_header(k)
            if "symbol" in nn or "ticker" in nn or "scrip" in nn or "stock" in nn:
                out["symbol"] = k
                break
    if not out["quantity"]:
        for k in keys:
            nn = _normalize_header(k)
            if "qty" in nn or "quantity" in nn or "balance" in nn or "shares" in nn:
                out["quantity"] = k
                break
    # If we have only one "value"-like column and no buy_value/closing_value, use it for value
    if not out["value"] and (out["buy_value"] or any("value" in _normalize_header(k) or "amount" in _normalize_header(k) for k in keys)):
        for k in keys:
            nn = _normalize_header(k)
            if ("value" in nn or "amount" in nn) and "buy" not in nn:
                out["value"] = k
                break
    return out


def _parse_float(row: dict[str, str], key: str | None) -> float | None:
    if not key or not row.get(key):
        return None
    try:
        return round(float(str(row.get(key, "")).replace(",", "")), 2)
    except ValueError:
        return None


def parse_csv(content: bytes, filename: str = "") -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parse CSV bytes. Returns (holdings, errors).
    Each holding: symbol, stock_name?, quantity, avg_cost?, buy_value?, closing_price?, value? (current), unrealized_pnl?.
    """
    holdings: list[dict[str, Any]] = []
    errors: list[str] = []
    try:
        text = content.decode("utf-8-sig").strip()
    except UnicodeDecodeError:
        try:
            text = content.decode("latin-1").strip()
        except Exception:
            errors.append("Could not decode file as UTF-8 or Latin-1.")
            return holdings, errors

    if not text:
        errors.append("File is empty.")
        return holdings, errors

    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        errors.append("No data rows found.")
        return holdings, errors

    first = rows[0]
    col = _detect_columns(first)
    sym_key = col.get("symbol") or col.get("stock_name")
    qty_key = col.get("quantity")
    if not sym_key or not qty_key:
        errors.append("Could not find symbol and quantity columns. Use headers like: symbol, quantity, avg_cost, value")
        return holdings, errors

    for i, row in enumerate(rows):
        sym = (row.get(col.get("symbol") or sym_key) or "").strip()
        if not sym and col.get("stock_name"):
            sym = (row.get(col["stock_name"]) or "").strip()
        if not sym:
            continue
        qty_str = (row.get(qty_key) or "").strip().replace(",", "")
        try:
            qty = int(float(qty_str)) if qty_str else 0
        except ValueError:
            errors.append(f"Row {i + 2}: invalid quantity '{qty_str}' for {sym}")
            continue
        if qty <= 0:
            continue

        stock_name = (row.get(col["stock_name"]) or "").strip() if col.get("stock_name") else None
        h: dict[str, Any] = {
            "symbol": sym,
            "quantity": qty,
            "avg_cost": _parse_float(row, col.get("avg_cost")),
            "buy_value": _parse_float(row, col.get("buy_value")),
            "closing_price": _parse_float(row, col.get("closing_price")),
            "value": _parse_float(row, col.get("value")),
            "unrealized_pnl": _parse_float(row, col.get("unrealized_pnl")),
        }
        if stock_name:
            h["stock_name"] = stock_name
        holdings.append(h)

    return holdings, errors


def parse_excel(content: bytes, filename: str = "") -> tuple[list[dict[str, Any]], list[str]]:
    """
    Parse first sheet of .xlsx. Returns (holdings, errors).
    Same holding shape as parse_csv.
    """
    if not HAS_EXCEL:
        return [], ["Excel support requires openpyxl. Install with: pip install openpyxl"]
    holdings: list[dict[str, Any]] = []
    errors: list[str] = []
    try:
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=_OPENPYXL_STYLE_WARNING)
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            errors.append("Workbook has no active sheet.")
            return holdings, errors
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        errors.append(f"Could not read Excel file: {e}")
        return holdings, errors

    if not rows:
        errors.append("Sheet is empty.")
        return holdings, errors

    # Find header row: first row where we detect symbol + quantity columns (broker sheets may have title rows)
    header_row_idx = None
    headers = []
    for idx, row in enumerate(rows):
        if not row:
            continue
        cand = [str(h or "").strip() for h in row]
        if not any(cand):
            continue
        header_dict = {h: "" for h in cand if h}
        col = _detect_columns(header_dict, use_all_keys=True)
        if (col.get("symbol") or col.get("stock_name")) and col.get("quantity"):
            header_row_idx = idx
            headers = cand
            break
    if header_row_idx is None or not headers:
        errors.append(
            "Could not find symbol and quantity columns in Excel. "
            "Expected headers like: Symbol / Trading Symbol, Quantity / Qty / Balance Qty; optionally Avg Cost / Value."
        )
        return holdings, errors

    data_rows = rows[header_row_idx + 1 :]
    if not data_rows:
        errors.append("No data rows below header.")
        return holdings, errors

    def row_to_dict(row: tuple) -> dict:
        r = list(row or [])
        while len(r) < len(headers):
            r.append(None)
        return dict(zip(headers, [v if v is None else str(v) for v in r[: len(headers)]]))

    row_dicts = [row_to_dict(row) for row in data_rows]
    header_dict = {h: "" for h in headers if h}
    col = _detect_columns(header_dict, use_all_keys=True)
    sym_key = col.get("symbol") or col.get("stock_name")
    qty_key = col.get("quantity")

    for i, row in enumerate(row_dicts):
        sym = (row.get(col.get("symbol") or sym_key) or "").strip()
        if not sym and col.get("stock_name"):
            sym = (row.get(col["stock_name"]) or "").strip()
        if not sym:
            continue
        qty_str = (row.get(qty_key) or "").strip().replace(",", "")
        try:
            qty = int(float(qty_str)) if qty_str else 0
        except ValueError:
            sheet_row = header_row_idx + i + 2
            errors.append(f"Row {sheet_row}: invalid quantity for {sym}")
            continue
        if qty <= 0:
            continue
        stock_name = (row.get(col["stock_name"]) or "").strip() if col.get("stock_name") else None
        h: dict[str, Any] = {
            "symbol": sym,
            "quantity": qty,
            "avg_cost": _parse_float(row, col.get("avg_cost")),
            "buy_value": _parse_float(row, col.get("buy_value")),
            "closing_price": _parse_float(row, col.get("closing_price")),
            "value": _parse_float(row, col.get("value")),
            "unrealized_pnl": _parse_float(row, col.get("unrealized_pnl")),
        }
        if stock_name:
            h["stock_name"] = stock_name
        holdings.append(h)
    return holdings, errors


def parse_portfolio_file(content: bytes, filename: str = "") -> tuple[list[dict[str, Any]], list[str]]:
    """
    Dispatch by extension: .csv -> parse_csv, .xlsx/.xls -> parse_excel.
    Returns (holdings, errors).
    """
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return parse_csv(content, filename)
    if fn.endswith(".xlsx") or fn.endswith(".xls"):
        return parse_excel(content, filename)
    return [], ["Unsupported file type. Use .csv or .xlsx"]